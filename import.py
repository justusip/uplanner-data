import datetime
import json
from collections import defaultdict

from openpyxl import load_workbook


def translate_term(term: str) -> str:
    s = term.lower()
    if s.endswith("sem 1"):
        return "s1"
    elif s.endswith("sem 2"):
        return "s2"
    elif s.endswith("sum sem"):
        return "ss"
    return "na"


def scrap(filename: str) -> []:
    workbook = load_workbook(filename=filename)
    worksheet = workbook.active
    courses = {}

    headers = [o.value.strip() for o in worksheet[1]]
    for raw_row in worksheet.iter_rows(values_only=True, min_row=2):
        row = {k: str(raw_row[i]).strip() for i, k in enumerate(headers)}

        term = translate_term(row["TERM"])
        course_code = row["COURSE CODE"]
        subclass = row["CLASS SECTION"]
        course_title = row["COURSE TITLE"]

        start_date = row["START DATE"]
        end_date = row["END DATE"]

        venue = row["VENUE"]
        start_time = row["START TIME"]
        end_time = row["END TIME"]

        index_name = f"{course_code}_{term}"
        print(f"Importing course {index_name}...")

        if index_name not in courses:
            courses[index_name] = {
                "code": course_code,
                "term": term,
                "title": course_title,
                "subclass": {}
            }

        if subclass not in courses[index_name]["subclass"]:
            courses[index_name]["subclass"][subclass] = []

        if start_date == "" or end_date == "":
            print(f"{index_name}: Start date or end date missing")
            continue
        if start_time == "" or end_time == "":
            print(f"{index_name}: Start time or end time missing")
            continue

        start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()

        weekday = None
        for i in range(7):
            if raw_row[headers.index("MON") + i] is not None:
                weekday = i
        if weekday is None:
            print(f"{index_name}: None of the weekday column is not empty.")
            continue

        start_time = datetime.datetime.strptime(start_time, "%H:%M").time()
        end_time = datetime.datetime.strptime(end_time, "%H:%M").time()

        for i in range((end_date - start_date).days + 1):
            day = start_date + datetime.timedelta(days=i)
            if day.weekday() == weekday:
                start = datetime.datetime.combine(day, start_time)
                end = datetime.datetime.combine(day, end_time)
                courses[index_name]["subclass"][subclass].append({
                    "from": start.isoformat(),
                    "to": end.isoformat(),
                    "venue": venue,
                })

        courses[index_name]["subclass"][subclass].sort(key=lambda o: o["from"], reverse=False)

    courses_list = list(courses.values())
    for course in courses_list:
        course["subclass"] = [
            {
                "name": sectionCode,
                "times": times
            } for sectionCode, times in course["subclass"].items()
        ]
    return courses_list


def export(institution: str, year: str, files: [str]):
    scraped_courses = []

    for file in files:
        scraped_courses.extend(scrap(file))

    d = defaultdict(dict)
    for c in scraped_courses:
        d[f"{c['code']}_{c['term']}"].update(c)
    cleaned_courses = list(d.values())
    cleaned_courses.sort(key=lambda c: c["code"])

    with open(f"{institution}_{year}.json", "w", encoding="utf8") as outFile:
        json.dump(cleaned_courses, outFile, sort_keys=True, indent=4, ensure_ascii=False)


# export("hku", "2021-2022", ["2021-22_class_timetable_00000000.xlsx", "2021-22_class_timetable_20220112.xlsx"])
export("hku", "2022-2023", ["2022-23_class_timetable_20220801.xlsx"])
